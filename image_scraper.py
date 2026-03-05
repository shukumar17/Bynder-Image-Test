"""Image Scraper and Screenshot Validator

This module scrapes image URLs from web pages, saves them to an Excel file,
and validates each image by taking screenshots.
"""

import logging
import os
import time
from typing import List, Tuple, Optional, Dict
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import requests
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from PIL import Image
import io
import shutil

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Import configuration
try:
    from config import AUTH_CONFIG, REQUEST_TIMEOUT, PAGE_LOAD_TIMEOUT
    logger.info("Configuration loaded from config.py")
except ImportError:
    logger.warning("config.py not found. Using default settings without authentication.")
    AUTH_CONFIG = {'username': '', 'password': ''}
    REQUEST_TIMEOUT = 30
    PAGE_LOAD_TIMEOUT = 30


class ImageScraper:
    """Scrapes images from URLs and validates them with screenshots."""

    def __init__(self, output_file: str = "Bynder.xlsx", username: str = None, password: str = None, max_workers: int = 5, initial_url: str = None):
        """
        Initialize the ImageScraper.

        Args:
            output_file: Path to the output Excel file
            username: Username for HTTP Basic Authentication (optional, uses config.py if not provided)
            password: Password for HTTP Basic Authentication (optional, uses config.py if not provided)
            max_workers: Maximum number of parallel workers for screenshot capture (default: 5)
            initial_url: URL to visit before starting tests (optional)
        """
        self.output_file = output_file
        self.driver: Optional[webdriver.Chrome] = None
        self.screenshot_dir = "screenshots"
        self.max_workers = max_workers
        self._driver_lock = threading.Lock()
        self.initial_url = initial_url
        
        # Set authentication credentials
        self.username = username or AUTH_CONFIG.get('username', '')
        self.password = password or AUTH_CONFIG.get('password', '')
        
        # Log authentication status
        if self.username and self.password:
            logger.info(f"Authentication enabled for user: {self.username}")
        else:
            logger.info("No authentication credentials provided")
        
        logger.info(f"Parallel execution enabled with {max_workers} workers")
        self._ensure_screenshot_directory()

    def _ensure_screenshot_directory(self) -> None:
        """Create screenshots directory if it doesn't exist."""
        if not os.path.exists(self.screenshot_dir):
            os.makedirs(self.screenshot_dir)
            logger.info(f"Created directory: {self.screenshot_dir}")

    def _create_driver(self, visit_initial: bool = False) -> webdriver.Chrome:
        """Create a new Selenium WebDriver instance for parallel execution.
        
        Args:
            visit_initial: If True, visit the initial URL (only for the first driver instance)
        """
        try:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--ignore-certificate-errors')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
            
            # Try to create driver without clearing cache (for parallel execution)
            try:
                driver_path = ChromeDriverManager(cache_valid_range=7).install()
                service = Service(driver_path)
                driver = webdriver.Chrome(service=service, options=chrome_options)
            except Exception:
                # Fallback to direct Chrome initialization
                driver = webdriver.Chrome(options=chrome_options)
            
            driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
            
            # Visit initial URL ONLY if explicitly requested (first time only)
            if visit_initial and self.initial_url:
                try:
                    logger.info(f"Visiting initial URL: {self.initial_url}")
                    authenticated_initial_url = self._get_authenticated_url(self.initial_url)
                    driver.get(authenticated_initial_url)
                    time.sleep(5)  # Wait for page to load
                    logger.info("Initial URL visited successfully")
                except Exception as init_error:
                    logger.warning(f"Could not visit initial URL: {init_error}")
            
            return driver
            
        except Exception as e:
            logger.error(f"Failed to create WebDriver: {e}")
            raise

    def _initialize_driver(self) -> None:
        """Initialize Selenium WebDriver with Chrome."""
        if self.driver is None:
            try:
                chrome_options = Options()
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--disable-dev-shm-usage')
                chrome_options.add_argument('--window-size=1920,1080')
                chrome_options.add_argument('--ignore-certificate-errors')
                chrome_options.add_argument('--disable-extensions')
                chrome_options.add_argument('--disable-blink-features=AutomationControlled')
                chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
                chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
                
                # Clear cache and force fresh ChromeDriver installation
                import shutil
                cache_paths = [
                    os.path.expanduser('~/.wdm'),
                    os.path.expanduser('~/.cache/selenium')
                ]
                for cache_path in cache_paths:
                    if os.path.exists(cache_path):
                        try:
                            shutil.rmtree(cache_path)
                            logger.info(f"Cleared cache: {cache_path}")
                        except Exception as cache_error:
                            logger.warning(f"Could not clear cache {cache_path}: {cache_error}")
                
                # Install ChromeDriver with explicit version matching
                try:
                    logger.info("Installing ChromeDriver...")
                    driver_path = ChromeDriverManager(cache_valid_range=0).install()
                    logger.info(f"ChromeDriver installed at: {driver_path}")
                    
                    # Verify the driver file exists and is executable
                    if not os.path.exists(driver_path):
                        raise FileNotFoundError(f"ChromeDriver not found at {driver_path}")
                    
                    service = Service(driver_path)
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    logger.info("WebDriver initialized successfully")
                    
                except Exception as driver_error:
                    logger.error(f"ChromeDriver installation failed: {driver_error}")
                    logger.info("Attempting alternative: using Chrome without webdriver-manager...")
                    
                    # Alternative: Try using Chrome directly without webdriver-manager
                    try:
                        self.driver = webdriver.Chrome(options=chrome_options)
                    except Exception as alt_error:
                        logger.error(f"Alternative method also failed: {alt_error}")
                        raise
                    
                self.driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
                
                # Visit initial URL if configured (for session establishment)
                if self.initial_url:
                    try:
                        logger.info(f"Visiting initial URL for session: {self.initial_url}")
                        authenticated_initial_url = self._get_authenticated_url(self.initial_url)
                        self.driver.get(authenticated_initial_url)
                        time.sleep(5)  # Wait for session to establish
                        logger.info("Initial URL visited, session established")
                    except Exception as init_error:
                        logger.warning(f"Could not visit initial URL: {init_error}")
                
            except Exception as e:
                logger.error(f"Failed to initialize WebDriver: {e}")
                logger.info("\n" + "="*60)
                logger.info("TROUBLESHOOTING STEPS:")
                logger.info("1. Ensure Google Chrome is installed and up to date")
                logger.info("2. Verify Python architecture matches Chrome (64-bit)")
                logger.info("3. Run in PowerShell with admin rights:")
                logger.info("   Remove-Item -Recurse -Force $env:USERPROFILE\\.wdm")
                logger.info("   Remove-Item -Recurse -Force $env:USERPROFILE\\.cache\\selenium")
                logger.info("4. Reinstall packages:")
                logger.info("   pip uninstall selenium webdriver-manager -y")
                logger.info("   pip install selenium webdriver-manager")
                logger.info("5. Check antivirus/firewall isn't blocking ChromeDriver")
                logger.info("="*60 + "\n")
                raise

    def _close_driver(self) -> None:
        """Close and cleanup WebDriver."""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("WebDriver closed successfully")
            except Exception as e:
                logger.error(f"Error closing WebDriver: {e}")
            finally:
                self.driver = None

    def _capture_screenshot_internal(self, index: int) -> Optional[str]:
        """Capture screenshot using existing driver session."""
        try:
            if self.driver:
                screenshot_filename = f"screenshot_{index}.png"
                screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                self.driver.save_screenshot(screenshot_path)
                if os.path.exists(screenshot_path) and os.path.getsize(screenshot_path) > 0:
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                    return screenshot_path
        except Exception as e:
            logger.warning(f"Could not capture fail scenario screenshot: {e}")
        return None

    def _capture_screenshot_for_fail(self, image_url: str, index: int) -> Optional[str]:
        """Capture screenshot for fail scenarios that need driver initialization."""
        try:
            if not self.driver:
                self._initialize_driver()
            
            # For URLs requiring authentication, modify URL to include credentials
            authenticated_url = image_url
            if self.username and self.password:
                parsed_url = urlparse(image_url)
                authenticated_url = f"{parsed_url.scheme}://{self.username}:{self.password}@{parsed_url.netloc}{parsed_url.path}"
                if parsed_url.query:
                    authenticated_url += f"?{parsed_url.query}"
                if parsed_url.fragment:
                    authenticated_url += f"#{parsed_url.fragment}"
            
            self.driver.get(authenticated_url)
            time.sleep(2)
            
            screenshot_filename = f"screenshot_{index}.png"
            screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
            self.driver.save_screenshot(screenshot_path)
            
            if os.path.exists(screenshot_path) and os.path.getsize(screenshot_path) > 0:
                logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                return screenshot_path
        except Exception as e:
            logger.warning(f"Could not capture fail scenario screenshot: {e}")
        return None

    def get_image_urls(self, page_url: str, use_selenium: bool = False) -> List[str]:
        """
        Extract image URLs ONLY from Picture tags, tile-swatches class, and nosto-container class elements.

        Args:
            page_url: The URL of the page to scrape
            use_selenium: If True, use Selenium to fetch page (maintains session from initial URL)

        Returns:
            List of image URLs found from <picture> tags and elements with class 'tile-swatches' and 'nosto-container'
        """
        image_urls = []
        
        try:
            # Validate URL
            parsed_url = urlparse(page_url)
            if not parsed_url.scheme or not parsed_url.netloc:
                logger.error(f"Invalid URL: {page_url}")
                return []

            logger.info(f"Fetching images from: {page_url}")
            logger.info("Targeting ONLY <picture> tags, 'tile-swatches' class, and 'nosto-container' class elements")
            
            # If initial_url is set, use Selenium to maintain session
            if self.initial_url or use_selenium:
                logger.info("Using Selenium to fetch page content (session-based)")
                if not self.driver:
                    self._initialize_driver()
                
                # Now visit the target page in the same session (initial URL already visited in _initialize_driver)
                authenticated_page_url = self._get_authenticated_url(page_url)
                logger.info(f"Opening target page in same session: {page_url}")
                self.driver.get(authenticated_page_url)
                time.sleep(3)  # Wait for page to load
                
                # Get page source from Selenium
                page_content = self.driver.page_source
                soup = BeautifulSoup(page_content, 'html.parser')
            else:
                # Use requests library (original behavior when no initial URL)
                logger.info("Using requests library to fetch page content")
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                
                # Add HTTP Basic Authentication if credentials are provided
                auth = None
                if self.username and self.password:
                    auth = HTTPBasicAuth(self.username, self.password)
                    logger.info(f"Using HTTP Basic Authentication for: {parsed_url.netloc}")
                
                response = requests.get(page_url, headers=headers, auth=auth, timeout=REQUEST_TIMEOUT)
                response.raise_for_status()
                
                # Parse HTML
                soup = BeautifulSoup(response.content, 'html.parser')
            
            # Strategy 1: Find all <picture> tags and extract images from them
            picture_tags = soup.find_all('picture')
            logger.info(f"Found {len(picture_tags)} <picture> tags")
            
            for picture in picture_tags:
                # Check for <source> tags within <picture>
                sources = picture.find_all('source')
                for source in sources:
                    # Get srcset or src attributes
                    srcset = source.get('srcset')
                    src = source.get('src')
                    
                    if srcset:
                        # srcset can contain multiple URLs separated by commas
                        # Format: "url1 descriptor1, url2 descriptor2, url3 descriptor3"
                        url_parts = srcset.split(',')
                        for url_part in url_parts:
                            # Each part might be "url descriptor" - take only the URL
                            # Strip whitespace first, then split and take first element
                            url_part_clean = url_part.strip()
                            if url_part_clean:
                                # Split by whitespace and take the first part (the URL)
                                url_clean = url_part_clean.split()[0] if ' ' in url_part_clean else url_part_clean
                                if url_clean:
                                    absolute_url = urljoin(page_url, url_clean)
                                    if absolute_url not in image_urls:
                                        image_urls.append(absolute_url)
                    elif src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
                
                # Also check for <img> tag within <picture>
                img = picture.find('img')
                if img:
                    srcset = img.get('srcset')
                    src = img.get('src') or img.get('data-src')
                    
                    if srcset:
                        # srcset can contain multiple URLs separated by commas
                        url_parts = srcset.split(',')
                        for url_part in url_parts:
                            url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                            if url_clean:
                                absolute_url = urljoin(page_url, url_clean)
                                if absolute_url not in image_urls:
                                    image_urls.append(absolute_url)
                    elif src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
            
            # Strategy 2: Find all elements with class 'tile-swatches' and extract images
            tile_swatches = soup.find_all(class_='tile-swatches')
            logger.info(f"Found {len(tile_swatches)} elements with class 'tile-swatches'")
            
            for tile in tile_swatches:
                # Find all img tags within tile-swatches
                img_tags = tile.find_all('img')
                for img in img_tags:
                    srcset = img.get('srcset')
                    src = img.get('src') or img.get('data-src')
                    
                    if srcset:
                        # srcset can contain multiple URLs separated by commas
                        url_parts = srcset.split(',')
                        for url_part in url_parts:
                            url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                            if url_clean:
                                absolute_url = urljoin(page_url, url_clean)
                                if absolute_url not in image_urls:
                                    image_urls.append(absolute_url)
                    elif src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
                
                # Also check for <picture> tags within tile-swatches
                pictures = tile.find_all('picture')
                for picture in pictures:
                    # Check for <source> tags
                    sources = picture.find_all('source')
                    for source in sources:
                        srcset = source.get('srcset')
                        src = source.get('src')
                        
                        if srcset:
                            url_parts = srcset.split(',')
                            for url_part in url_parts:
                                url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                                if url_clean:
                                    absolute_url = urljoin(page_url, url_clean)
                                    if absolute_url not in image_urls:
                                        image_urls.append(absolute_url)
                        elif src:
                            absolute_url = urljoin(page_url, src)
                            if absolute_url not in image_urls:
                                image_urls.append(absolute_url)
                    
                    # Check for <img> within <picture>
                    img = picture.find('img')
                    if img:
                        srcset = img.get('srcset')
                        src = img.get('src') or img.get('data-src')
                        
                        if srcset:
                            url_parts = srcset.split(',')
                            for url_part in url_parts:
                                url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                                if url_clean:
                                    absolute_url = urljoin(page_url, url_clean)
                                    if absolute_url not in image_urls:
                                        image_urls.append(absolute_url)
                        elif src:
                            absolute_url = urljoin(page_url, src)
                            if absolute_url not in image_urls:
                                image_urls.append(absolute_url)
            
            # Strategy 3: Find all elements with class 'nosto-container' and extract images
            nosto_containers = soup.find_all(class_='nosto-container')
            logger.info(f"Found {len(nosto_containers)} elements with class 'nosto-container'")
            
            for nosto in nosto_containers:
                # Find all img tags within nosto-container
                img_tags = nosto.find_all('img')
                for img in img_tags:
                    srcset = img.get('srcset')
                    src = img.get('src') or img.get('data-src')
                    
                    if srcset:
                        # srcset can contain multiple URLs separated by commas
                        url_parts = srcset.split(',')
                        for url_part in url_parts:
                            url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                            if url_clean:
                                absolute_url = urljoin(page_url, url_clean)
                                if absolute_url not in image_urls:
                                    image_urls.append(absolute_url)
                    elif src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
                
                # Also check for <picture> tags within nosto-container
                pictures = nosto.find_all('picture')
                for picture in pictures:
                    # Check for <source> tags
                    sources = picture.find_all('source')
                    for source in sources:
                        srcset = source.get('srcset')
                        src = source.get('src')
                        
                        if srcset:
                            url_parts = srcset.split(',')
                            for url_part in url_parts:
                                url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                                if url_clean:
                                    absolute_url = urljoin(page_url, url_clean)
                                    if absolute_url not in image_urls:
                                        image_urls.append(absolute_url)
                        elif src:
                            absolute_url = urljoin(page_url, src)
                            if absolute_url not in image_urls:
                                image_urls.append(absolute_url)
                    
                    # Check for <img> within <picture>
                    img = picture.find('img')
                    if img:
                        srcset = img.get('srcset')
                        src = img.get('src') or img.get('data-src')
                        
                        if srcset:
                            url_parts = srcset.split(',')
                            for url_part in url_parts:
                                url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                                if url_clean:
                                    absolute_url = urljoin(page_url, url_clean)
                                    if absolute_url not in image_urls:
                                        image_urls.append(absolute_url)
                        elif src:
                            absolute_url = urljoin(page_url, src)
                            if absolute_url not in image_urls:
                                image_urls.append(absolute_url)
            
            # Strategy 4: Find all <video> tags and extract video source URLs
            video_tags = soup.find_all('video')
            logger.info(f"Found {len(video_tags)} <video> tags")
            
            for video in video_tags:
                # Check for <source> tags within <video>
                sources = video.find_all('source')
                for source in sources:
                    src = source.get('src')
                    if src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
                            logger.info(f"Found video source: {absolute_url}")
                
                # Also check for src attribute directly on video tag
                video_src = video.get('src')
                if video_src:
                    absolute_url = urljoin(page_url, video_src)
                    if absolute_url not in image_urls:
                        image_urls.append(absolute_url)
                        logger.info(f"Found video source: {absolute_url}")
            
            # Strategy 5: Find all elements with class 'product-tile' or 'swiper-slide' and extract images
            product_tiles = soup.find_all(class_=['product-tile', 'swiper-slide'])
            logger.info(f"Found {len(product_tiles)} elements with class 'product-tile' or 'swiper-slide'")
            
            for tile in product_tiles:
                # Find all img tags within product-tile/swiper-slide
                img_tags = tile.find_all('img')
                for img in img_tags:
                    srcset = img.get('srcset')
                    src = img.get('src') or img.get('data-src')
                    
                    if srcset:
                        # srcset can contain multiple URLs separated by commas
                        url_parts = srcset.split(',')
                        for url_part in url_parts:
                            url_clean = url_part.strip().split()[0] if url_part.strip() else ''
                            if url_clean:
                                absolute_url = urljoin(page_url, url_clean)
                                if absolute_url not in image_urls:
                                    image_urls.append(absolute_url)
                    elif src:
                        absolute_url = urljoin(page_url, src)
                        if absolute_url not in image_urls:
                            image_urls.append(absolute_url)
            
            logger.info(f"Total: Found {len(image_urls)} unique image/video URLs from <picture> tags, 'tile-swatches' class, 'nosto-container' class, 'product-tile' class, 'swiper-slide' class, and <video> tags")
            
            # Filter to only include HTTPS URLs
            https_urls = [url for url in image_urls if url.startswith('https://')]
            filtered_count = len(image_urls) - len(https_urls)
            
            if filtered_count > 0:
                logger.info(f"Filtered out {filtered_count} non-HTTPS URLs")
            
            logger.info(f"Final count: {len(https_urls)} HTTPS image URLs")
            return https_urls
            
        except requests.RequestException as e:
            logger.error(f"Error fetching page {page_url}: {e}")
            return []  # Return empty list on error
        except Exception as e:
            logger.error(f"Unexpected error processing {page_url}: {e}")
            return []  # Return empty list on error

    def _take_screenshot_worker(self, image_url: str, index: int, is_first: bool = False) -> Tuple[str, Optional[str], str, str]:
        """
        Worker function for parallel screenshot capture - ONLY captures screenshots for failed cases.
        
        Args:
            image_url: URL of the image to screenshot
            index: Index for naming the screenshot file
            is_first: If True, this is the first worker and should visit initial URL
        
        Returns:
            Tuple of (image_url, screenshot_path, status, failure_reason)
        """
        driver = None
        try:
            # Check if URL is a video file (mp4, webm, ogg, etc.)
            is_video = any(ext in image_url.lower() for ext in ['.mp4', '.webm', '.ogg', '.mov', '.avi'])
            
            # For video URLs: Pass if from Bynder, Fail if from Amplience
            if is_video:
                if 'bynder.com' in image_url.lower():
                    logger.info(f"Video from Bynder (Pass): {image_url}")
                    return image_url, None, "Pass", ""
                elif 'amplience' in image_url.lower():
                    logger.warning(f"Video from Amplience (Fail): {image_url}")
                    # Create driver for fail scenario screenshot
                    driver = self._create_driver(visit_initial=is_first)
                    authenticated_url = self._get_authenticated_url(image_url)
                    
                    # Open in new tab if initial URL was visited
                    if is_first and self.initial_url:
                        driver.execute_script("window.open('');")
                        driver.switch_to.window(driver.window_handles[-1])
                    
                    driver.get(authenticated_url)
                    time.sleep(2)
                    
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    
                    return image_url, screenshot_path if os.path.exists(screenshot_path) else None, "Fail", "Video from Amplience"
            
            # Check if URL contains 'amplience' - fail if found (for images)
            if 'amplience' in image_url.lower():
                logger.warning(f"Image Coming from Amplience: {image_url}")
                # Create driver for fail scenario screenshot
                driver = self._create_driver(visit_initial=is_first)
                authenticated_url = self._get_authenticated_url(image_url)
                
                # Open in new tab if initial URL was visited
                if is_first and self.initial_url:
                    driver.execute_script("window.open('');")
                    driver.switch_to.window(driver.window_handles[-1])
                
                driver.get(authenticated_url)
                time.sleep(2)
                
                screenshot_filename = f"screenshot_{index}.png"
                screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                driver.save_screenshot(screenshot_path)
                
                return image_url, screenshot_path if os.path.exists(screenshot_path) else None, "Fail", "Amplience found in URL"
            
            # Create a new driver instance for this worker
            driver = self._create_driver(visit_initial=is_first)
            
            logger.info(f"Validating image URL: {image_url}")
            
            # Get authenticated URL
            authenticated_url = self._get_authenticated_url(image_url)
            
            # Open in new tab if initial URL was visited, otherwise use current tab
            if is_first and self.initial_url:
                driver.execute_script("window.open('');")
                driver.switch_to.window(driver.window_handles[-1])
            
            # Navigate to the image URL
            driver.get(authenticated_url)
            time.sleep(2)  # Wait for image to load
            
            # Check if page has actual image content or error messages
            try:
                page_source = driver.page_source
                page_source_lower = page_source.lower()
                
                # Check for specific error messages that should fail the test
                if 'no results for url found: consider case sensitivity' in page_source_lower:
                    logger.warning(f"Error message found: 'No results for url found: consider case sensitivity' at {image_url}")
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                    return image_url, screenshot_path, "Fail", "No results for url found: consider case sensitivity"
                
                if 'bad request' in page_source_lower:
                    logger.warning(f"Error message found: 'Bad Request' at {image_url}")
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                    return image_url, screenshot_path, "Fail", "Bad Request"
                
                # Check if the page shows an error or no image
                if 'not found' in page_source_lower or '404' in page_source_lower or 'error' in page_source_lower:
                    logger.warning(f"No valid image found at URL: {image_url}")
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                    return image_url, screenshot_path, "Fail", "No image found at URL"
            except Exception as check_error:
                logger.warning(f"Could not verify page content: {check_error}")
            
            # If we reach here, the image loaded successfully - NO SCREENSHOT for pass cases
            logger.info(f"Image validated successfully (Pass): {image_url}")
            return image_url, None, "Pass", ""
                
        except (WebDriverException, TimeoutException) as e:
            logger.error(f"WebDriver error for {image_url}: {e}")
            screenshot_path = None
            if driver:
                try:
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                except:
                    pass
            return image_url, screenshot_path, "Fail", f"WebDriver error: {str(e)[:50]}"
        except Exception as e:
            logger.error(f"Unexpected error validating {image_url}: {e}")
            # Try to capture screenshot for unexpected errors
            screenshot_path = None
            if driver:
                try:
                    screenshot_filename = f"screenshot_{index}.png"
                    screenshot_path = os.path.join(self.screenshot_dir, screenshot_filename)
                    driver.save_screenshot(screenshot_path)
                    logger.info(f"Fail scenario screenshot saved: {screenshot_path}")
                except:
                    pass
            return image_url, screenshot_path, "Fail", f"Error: {str(e)[:50]}"
        finally:
            # Close the driver instance
            if driver:
                try:
                    driver.quit()
                except:
                    pass

    def _get_authenticated_url(self, image_url: str) -> str:
        """Generate authenticated URL if credentials are provided."""
        authenticated_url = image_url
        if self.username and self.password:
            parsed_url = urlparse(image_url)
            authenticated_url = f"{parsed_url.scheme}://{self.username}:{self.password}@{parsed_url.netloc}{parsed_url.path}"
            if parsed_url.query:
                authenticated_url += f"?{parsed_url.query}"
            if parsed_url.fragment:
                authenticated_url += f"#{parsed_url.fragment}"
        return authenticated_url

    def take_screenshot(self, image_url: str, index: int) -> Tuple[Optional[str], str, str]:
        """
        Take a screenshot of an image URL and validate it.
        This is kept for backward compatibility but uses the worker function.

        Args:
            image_url: URL of the image to screenshot
            index: Index for naming the screenshot file

        Returns:
            Tuple of (screenshot_path, status, failure_reason)
        """
        _, screenshot_path, status, failure_reason = self._take_screenshot_worker(image_url, index)
        return screenshot_path, status, failure_reason

    def create_excel_report_for_page(self, page_url: str, page_folder: str, pass_filename: str, fail_filename: str) -> None:
        """
        Create separate Excel reports for Pass and Fail scenarios for a single page URL.

        Args:
            page_url: The page URL to process
            page_folder: Folder path to store the Excel files
            pass_filename: Name of the Excel file for Pass scenarios
            fail_filename: Name of the Excel file for Fail scenarios
        """
        try:
            # Create workbooks for Pass and Fail scenarios
            wb_pass = Workbook()
            ws_pass = wb_pass.active
            ws_pass.title = "Pass Scenarios"
            
            wb_fail = Workbook()
            ws_fail = wb_fail.active
            ws_fail.title = "Fail Scenarios"
            
            # Set headers for Pass workbook
            ws_pass['A1'] = "Page URL"
            ws_pass['B1'] = "Image URL"
            ws_pass['C1'] = "Screenshot"
            ws_pass['D1'] = "Failure Reason"
            
            # Set column widths for Pass workbook
            ws_pass.column_dimensions['A'].width = 50
            ws_pass.column_dimensions['B'].width = 60
            ws_pass.column_dimensions['C'].width = 30
            ws_pass.column_dimensions['D'].width = 40
            
            # Set headers for Fail workbook
            ws_fail['A1'] = "Page URL"
            ws_fail['B1'] = "Image URL"
            ws_fail['C1'] = "Screenshot"
            ws_fail['D1'] = "Failure Reason"
            
            # Set column widths for Fail workbook
            ws_fail.column_dimensions['A'].width = 50
            ws_fail.column_dimensions['B'].width = 60
            ws_fail.column_dimensions['C'].width = 30
            ws_fail.column_dimensions['D'].width = 40
            
            pass_row = 2
            fail_row = 2
            image_index = 1
            
            logger.info(f"Processing page: {page_url}")
            
            # Get all image URLs from the page
            image_urls = self.get_image_urls(page_url)
            
            # Fail test if no images found on the page
            if not image_urls:
                logger.warning(f"No images found on {page_url}")
                ws_fail[f'A{fail_row}'] = page_url
                ws_fail[f'B{fail_row}'] = "N/A"
                ws_fail[f'D{fail_row}'] = "No images found on page"
                fail_row += 1
            else:
                # Process images in parallel using ThreadPoolExecutor
                logger.info(f"Processing {len(image_urls)} images in parallel with {self.max_workers} workers")
                print(f"Processing {len(image_urls)} images in parallel with {self.max_workers} workers...")
                
                # Create tasks for parallel execution
                screenshot_results = {}
                
                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    # Submit all screenshot tasks - first task visits initial URL
                    future_to_url = {
                        executor.submit(self._take_screenshot_worker, img_url, image_index + idx, idx == 0): (img_url, image_index + idx)
                        for idx, img_url in enumerate(image_urls)
                    }
                    
                    # Collect results as they complete
                    completed = 0
                    total = len(future_to_url)
                    for future in as_completed(future_to_url):
                        try:
                            img_url, screenshot_path, status, failure_reason = future.result()
                            screenshot_results[img_url] = (screenshot_path, status, failure_reason)
                            completed += 1
                            print(f"Progress: {completed}/{total} screenshots completed")
                        except Exception as e:
                            img_url, _ = future_to_url[future]
                            logger.error(f"Error processing {img_url}: {e}")
                            screenshot_results[img_url] = (None, "Fail", f"Error: {str(e)[:50]}")
                            completed += 1
                
                # Process results in original order and add to Excel
                for img_url in image_urls:
                    screenshot_path, status, failure_reason = screenshot_results.get(img_url, (None, "Fail", "Unknown error"))
                    
                    # Determine which workbook to use
                    if status == "Pass":
                        ws = ws_pass
                        row = pass_row
                        pass_row += 1
                    else:
                        ws = ws_fail
                        row = fail_row
                        fail_row += 1
                    
                    # Add page URL in first column
                    ws[f'A{row}'] = page_url
                    
                    # Add image URL in second column
                    ws[f'B{row}'] = img_url
                    
                    # Add screenshot to Excel if successful
                    if screenshot_path and os.path.exists(screenshot_path):
                        try:
                            # Resize image for Excel
                            img = Image.open(screenshot_path)
                            img.thumbnail((200, 200))
                            
                            # Save resized image
                            resized_path = screenshot_path.replace('.png', '_thumb.png')
                            img.save(resized_path)
                            
                            # Add to Excel (column C)
                            xl_img = XLImage(resized_path)
                            xl_img.width = 200
                            xl_img.height = 200
                            ws.add_image(xl_img, f'C{row}')
                            
                            # Set row height
                            ws.row_dimensions[row].height = 150
                            
                        except Exception as e:
                            logger.error(f"Error adding screenshot to Excel: {e}")
                    
                    # Add failure reason if any (column D)
                    if failure_reason:
                        ws[f'D{row}'] = failure_reason
                    
                    image_index += 1
            
            # Create full paths for Excel files in page folder
            pass_filepath = os.path.join(page_folder, pass_filename)
            fail_filepath = os.path.join(page_folder, fail_filename)
            
            # Save both workbooks
            wb_pass.save(pass_filepath)
            logger.info(f"Pass scenarios Excel report saved: {pass_filepath}")
            
            wb_fail.save(fail_filepath)
            logger.info(f"Fail scenarios Excel report saved: {fail_filepath}")
            
            # Delete screenshots folder after saving to Excel
            try:
                if os.path.exists(self.screenshot_dir):
                    shutil.rmtree(self.screenshot_dir)
                    logger.info(f"Deleted screenshots folder: {self.screenshot_dir}")
                    # Recreate the directory for next page
                    self._ensure_screenshot_directory()
            except Exception as cleanup_error:
                logger.warning(f"Could not delete screenshots folder: {cleanup_error}")
            
        except Exception as e:
            logger.error(f"Error creating Excel reports: {e}")
            raise

    def run_with_page_wise_excel(self, urls: List[str]) -> None:
        """
        Main execution method - creates separate Pass and Fail Excel files for each page URL.

        Args:
            urls: List of page URLs to process
        """
        if not urls:
            logger.error("No URLs provided")
            return
        
        logger.info(f"Starting image scraping for {len(urls)} URL(s)")
        logger.info("Creating separate Pass and Fail Excel files for each page URL")
        
        # Create Tests directory if it doesn't exist
        tests_dir = "Tests"
        if not os.path.exists(tests_dir):
            os.makedirs(tests_dir)
            logger.info(f"Created Tests directory: {tests_dir}")
            print(f"Created Tests directory: {tests_dir}")
        
        try:
            for idx, page_url in enumerate(urls, 1):
                # Ask user for Excel filename base for this page
                print("\n" + "="*60)
                print(f"Processing Page {idx}/{len(urls)}: {page_url}")
                print("="*60)
                
                # Generate default filename base
                default_base = f"Bynder_Page_{idx}"
                
                # Ask user for custom filename base
                user_filename = input(f"Enter base filename for this page (press Enter for '{default_base}'): ").strip()
                
                # Use default if user didn't provide a name
                if not user_filename:
                    base_filename = default_base
                else:
                    # Remove .xlsx extension if provided
                    if user_filename.endswith('.xlsx'):
                        base_filename = user_filename[:-5]
                    else:
                        base_filename = user_filename
                
                # Create folder for this page inside Tests directory
                page_folder = os.path.join(tests_dir, base_filename)
                if not os.path.exists(page_folder):
                    os.makedirs(page_folder)
                    logger.info(f"Created folder: {page_folder}")
                    print(f"Created folder: {page_folder}")
                
                # Create Pass and Fail filenames
                pass_filename = f"{base_filename}_Pass.xlsx"
                fail_filename = f"{base_filename}_Fail.xlsx"
                
                logger.info(f"Creating Excel reports in folder '{page_folder}': {pass_filename} and {fail_filename}")
                print(f"Creating Excel reports in folder '{page_folder}': {pass_filename} and {fail_filename}")
                
                # Create Excel reports for this page
                self.create_excel_report_for_page(page_url, page_folder, pass_filename, fail_filename)
                
                print(f"✓ Pass scenarios saved to: {os.path.join(page_folder, pass_filename)}")
                print(f"✓ Fail scenarios saved to: {os.path.join(page_folder, fail_filename)}")
            
            logger.info("Image scraping completed successfully")
            print("\n" + "="*60)
            print("All page tests completed!")
            print(f"All results saved in '{tests_dir}' directory")
            print("="*60)
            
        except Exception as e:
            logger.error(f"Error during execution: {e}")
        finally:
            self._close_driver()


def main() -> None:
    """Main entry point for the script."""
    print("=" * 60)
    print("Image Scraper and Screenshot Validator")
    print("=" * 60)
    print()
    print("This tool will:")
    print("1. Visit each page URL you provide")
    print("2. Find image URLs from <picture> tags, 'tile-swatches' class, and 'nosto-container' class")
    print("3. Find video URLs from <video> tags")
    print("4. Filter to only include HTTPS URLs")
    print("5. Create separate Pass and Fail Excel files for each page URL")
    print("6. Validate each image/video IN PARALLEL (faster processing)")
    print("7. Take screenshots ONLY for FAILED cases")
    print("8. Add screenshots to Fail Excel report")
    print("9. Delete screenshots folder after saving to Excel")
    print("10. Fail tests for:")
    print("   - Non-HTTPS URLs")
    print("   - Amplience URLs (images and videos)")
    print("   - 'No results for url found: consider case sensitivity'")
    print("   - 'Bad Request' errors")
    print("   - Pages with no images/videos")
    print("11. Pass tests for:")
    print("   - Videos from Bynder (marcjacobs.bynder.com)")
    print()
    
    # Get URLs from user input
    user_input = input("Enter page URL(s) separated by commas: ").strip()
    
    if not user_input:
        print("No URLs provided. Exiting.")
        return
    
    # Parse URLs
    urls = [url.strip() for url in user_input.split(',') if url.strip()]
    
    if not urls:
        print("No valid URLs provided. Exiting.")
        return
    
    # Ask if URL is password protected
    print("\n" + "="*60)
    is_protected = input("Is the URL password protected? (yes/no): ").strip().lower()
    
    username = None
    password = None
    
    if is_protected in ['yes', 'y']:
        print("\nPlease enter authentication credentials:")
        username = input("Username: ").strip()
        password = input("Password: ").strip()
        
        if username and password:
            print(f"✓ Authentication credentials set for user: {username}")
        else:
            print("⚠ Warning: Empty credentials provided. Proceeding without authentication.")
            username = None
            password = None
    else:
        print("✓ Proceeding without authentication")
    
    # Ask for initial URL to visit before tests
    print("\n" + "="*60)
    visit_initial = input("Do you want to visit a URL before tests start? (yes/no): ").strip().lower()
    
    initial_url = None
    if visit_initial in ['yes', 'y']:
        initial_url = input("Enter the URL to visit before tests start: ").strip()
        if initial_url:
            print(f"✓ Will visit {initial_url} before each test starts")
            print("✓ Tests will run in new tabs")
        else:
            print("⚠ Warning: Empty URL provided. Proceeding without initial URL.")
            initial_url = None
    else:
        print("✓ Proceeding without initial URL visit")
    
    # Set parallel execution to 10 workers (no user prompt needed)
    max_workers = 10
    print("="*60)
    print(f"✓ Using {max_workers} parallel workers for faster processing")
    print("✓ Screenshots will be captured ONLY for failed cases")
    print("="*60)
    print(f"\nProcessing {len(urls)} page URL(s)...")
    print("You will be asked to provide base filename for each page (Pass and Fail files will be created).\n")
    
    # Create scraper with user-provided credentials, parallel workers, and initial URL
    scraper = ImageScraper(username=username, password=password, max_workers=max_workers, initial_url=initial_url)
    scraper.run_with_page_wise_excel(urls)
    
    print("\n" + "=" * 60)
    print("Process completed! Check Pass and Fail Excel files for results.")
    print("=" * 60)


if __name__ == "__main__":
    main()